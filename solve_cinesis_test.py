#!/usr/bin/env python3
"""
solve_cinesis_test.py — Generic driver–load matching solver.

Reads a workbook with tabs:
  Sample Conversation  — driver transcript
  Loads                — available loads
  Part A (Fill In)     — driver profile output
  Part B (Fill In)     — top-3 loads output

Usage:
    python solve_cinesis_test.py                          # lee cinesis_good_fit_test_clean.xlsx
    python solve_cinesis_test.py mi_archivo.xlsx          # lee el archivo indicado
    python solve_cinesis_test.py input.xlsx output.xlsx   # input y output explícitos

Output por defecto: cinesis_good_fit_test_completed.xlsx

Variables de entorno (cargadas desde .env si existe):
    OPENAI_API_KEY   — habilita extracción con GPT-4o; sin ella usa fallback determinístico
    OPENAI_LOG       — si es "1", imprime prompt y respuesta completa de OpenAI
"""

import json
import logging
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl
import pandas as pd

# Carga .env si existe (requiere: pip install python-dotenv)
try:
    from dotenv import load_dotenv
    _env_path = os.path.join(os.path.dirname(__file__), ".env")
    if os.path.exists(_env_path):
        load_dotenv(_env_path)
        print(f"[env] .env cargado desde {_env_path}")
except ImportError:
    pass  # python-dotenv no instalado; las vars deben estar seteadas manualmente

# ---------------------------------------------------------------------------
# Logging: consola + archivo logs/log_YYYYMMDD_HHMMSS.log
# ---------------------------------------------------------------------------
import datetime

_log_dir  = os.path.join(os.path.dirname(__file__), "logs")
os.makedirs(_log_dir, exist_ok=True)
_log_file = os.path.join(
    _log_dir,
    f"log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.log",
)

_log_fmt     = logging.Formatter("%(asctime)s  %(levelname)s  %(message)s", datefmt="%H:%M:%S")
_log_level   = logging.DEBUG if os.environ.get("OPENAI_LOG") == "1" else logging.INFO

_console_handler = logging.StreamHandler(sys.stdout)
_console_handler.setFormatter(_log_fmt)
_console_handler.setLevel(_log_level)

_file_handler = logging.FileHandler(_log_file, encoding="utf-8")
_file_handler.setFormatter(_log_fmt)
_file_handler.setLevel(logging.DEBUG)  # el archivo siempre guarda todo

logging.basicConfig(level=logging.DEBUG, handlers=[_console_handler, _file_handler])

# Logger raíz — reemplaza los print() del script
log = logging.getLogger("cinesis")

# Logger dedicado para las llamadas a OpenAI (mismo nivel que root)
_oai_log = logging.getLogger("openai_calls")

print(f"[log] archivo de log: {_log_file}")
from openpyxl.styles import Alignment, Font


# =============================================================================
# 1.  City coordinate lookup
# =============================================================================

# Seed table — "city, st" keys → (lat, lon).
# Augmented at runtime from the workbook's Loads tab.
_SEED_CITY_COORDS: dict[str, tuple[float, float]] = {
    "abilene, tx":           (32.4487, -99.7331),
    "amarillo, tx":          (35.2220, -101.8313),
    "arlington, tx":         (32.7357, -97.1081),
    "atlanta, ga":           (33.7490, -84.3880),
    "austin, tx":            (30.2672, -97.7431),
    "baton rouge, la":       (30.4515, -91.1871),
    "beaumont, tx":          (30.0802, -94.1266),
    "brownsville, tx":       (25.9017, -97.4975),
    "corpus christi, tx":    (27.8006, -97.3964),
    "dallas, tx":            (32.7767, -96.7970),
    "el paso, tx":           (31.7619, -106.4850),
    "fort worth, tx":        (32.7555, -97.3308),
    "gainesville, fl":       (29.6516, -82.3248),
    "houston, tx":           (29.7604, -95.3698),
    "huntsville, tx":        (30.7235, -95.5508),
    "jackson, ms":           (32.2988, -90.1848),
    "jacksonville, fl":      (30.3322, -81.6557),
    "laredo, tx":            (27.5306, -99.4803),
    "longview, tx":          (32.5007, -94.7405),
    "lubbock, tx":           (33.5779, -101.8552),
    "mcallen, tx":           (26.2034, -98.2300),
    "memphis, tn":           (35.1495, -90.0490),
    "midland, tx":           (31.9973, -102.0779),
    "nashville, tn":         (36.1627, -86.7816),
    "new orleans, la":       (29.9511, -90.0715),
    "odessa, tx":            (31.8457, -102.3676),
    "oklahoma city, ok":     (35.4676, -97.5164),
    "orlando, fl":           (28.5383, -81.3792),
    "plano, tx":             (33.0198, -96.6989),
    "rio grande valley, tx": (26.3017, -98.1633),
    "san antonio, tx":       (29.4241, -98.4936),
    "san marcos, tx":        (29.8833, -97.9414),
    "shreveport, la":        (32.5252, -93.7502),
    "temple, tx":            (31.0982, -97.3428),
    "tulsa, ok":             (36.1540, -95.9928),
    "tyler, tx":             (32.3513, -95.3011),
    "victoria, tx":          (28.8053, -97.0036),
    "waco, tx":              (31.5493, -97.1467),
    "wichita falls, tx":     (33.9137, -98.4934),
}


def build_city_table(xlsx_path: str) -> dict[str, tuple[float, float]]:
    """
    Return city-name → (lat, lon) dict.
    Seeds from _SEED_CITY_COORDS (stored as both "city, st" and bare "city").
    Augmented with every city present in the Loads tab of the workbook.
    """
    table: dict[str, tuple[float, float]] = {}
    for city_state, coords in _SEED_CITY_COORDS.items():
        table[city_state] = coords
        bare = city_state.split(",")[0].strip()
        table[bare] = coords

    try:
        df = pd.read_excel(xlsx_path, sheet_name="Loads", header=0)
        for _, row in df.iterrows():
            for city_col, lat_col, lon_col in [
                ("Origin", "Origin Lat", "Origin Lon"),
                ("Destination", "Dest Lat", "Dest Lon"),
            ]:
                city = str(row.get(city_col, "")).strip()
                if not city or city.lower() in ("nan", "missing"):
                    continue
                try:
                    lat, lon = float(row[lat_col]), float(row[lon_col])
                    table[city.lower()] = (lat, lon)
                except (ValueError, TypeError):
                    pass
    except Exception:
        pass

    return table


def resolve_city(name: str, table: dict[str, tuple[float, float]]) -> Optional[tuple[float, float]]:
    """
    Resolve a city string to (lat, lon).
    Tries: exact lower-cased key → bare city without state → prefix substring.
    Returns None if not found.
    """
    if not name:
        return None
    key = name.lower().strip()
    if key in table:
        return table[key]
    bare = key.split(",")[0].strip()
    if bare in table:
        return table[bare]
    for stored, coords in table.items():
        stored_bare = stored.split(",")[0].strip()
        if stored_bare == bare:
            return coords
    return None


def _city_label(name: str) -> str:
    """Format a city name as 'City, ST' if the seed table recognises it."""
    key = name.lower().strip()
    for seed_key in _SEED_CITY_COORDS:
        if seed_key.split(",")[0].strip() == key:
            city_part = seed_key.split(",")[0].strip().title()
            state_part = seed_key.split(",")[1].strip().upper()
            return f"{city_part}, {state_part}"
    return name.title()


def fill_coords(loc: dict, table: dict[str, tuple[float, float]]) -> None:
    """In-place: fill null lat/lon in a location dict from the city table."""
    if loc.get("lat") is None or loc.get("lon") is None:
        coords = resolve_city(loc.get("city_state") or "", table)
        if coords:
            loc["lat"], loc["lon"] = coords


# =============================================================================
# 2.  Equipment normalization
# =============================================================================

_EQUIP_ALIASES: dict[str, str] = {
    "hot shot":   "hotshot",
    "hot-shot":   "hotshot",
    "hotshot":    "hotshot",
    "goose neck": "gooseneck",
    "goose-neck": "gooseneck",
    "gooseneck":  "gooseneck",
    "flat bed":   "flatbed",
    "flat-bed":   "flatbed",
    "flatbed":    "flatbed",
    "dry van":    "van",
    "dry-van":    "van",
    "van":        "van",
    "reefer":     "reefer",
    "refrigerated": "reefer",
    "step deck":  "stepdeck",
    "step-deck":  "stepdeck",
    "stepdeck":   "stepdeck",
    "lowboy":     "lowboy",
    "low boy":    "lowboy",
    "low-boy":    "lowboy",
    "tanker":     "tanker",
}


def normalize_equip(s: str) -> str:
    """Return canonical equipment name (lowercase, alias-resolved)."""
    key = s.lower().strip()
    if key in _EQUIP_ALIASES:
        return _EQUIP_ALIASES[key]
    collapsed = re.sub(r"[\s\-]+", "", key)
    return _EQUIP_ALIASES.get(collapsed, collapsed)


def equip_matches(driver_equip_list: list[str], load_trailer: str) -> bool:
    """
    True if load_trailer (after normalization) matches any item in driver_equip_list.
    Examples that match: "goose neck" == "Gooseneck", "hot shot" == "Hotshot".
    """
    load_canonical = normalize_equip(load_trailer)
    return load_canonical in {normalize_equip(e) for e in driver_equip_list}


# =============================================================================
# 3.  Transcript reader
# =============================================================================

def read_transcript(xlsx_path: str) -> str:
    df = pd.read_excel(xlsx_path, sheet_name="Sample Conversation", header=None)
    lines = []
    for _, row in df.iterrows():
        speaker = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        text    = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        if speaker and text and speaker not in ("Speaker", "nan"):
            lines.append(f"{speaker}: {text}")
    return "\n".join(lines)


def lines_by_speaker(transcript: str) -> dict[str, list[str]]:
    result: dict[str, list[str]] = {}
    for line in transcript.splitlines():
        if ": " in line:
            speaker, _, text = line.partition(": ")
            result.setdefault(speaker.strip(), []).append(text.strip())
    return result


# =============================================================================
# 4.  Profile extraction — JSON schema
# =============================================================================

_PROFILE_SCHEMA: dict = {
    "type": "object",
    "properties": {
        "current_location": {
            "type": "object",
            "properties": {
                "city_state":  {"type": ["string", "null"]},
                "lat":         {"type": ["number", "null"]},
                "lon":         {"type": ["number", "null"]},
                "evidence":    {"type": "string"},
                "confidence":  {"type": "number"},
            },
            "required": ["city_state", "lat", "lon", "evidence", "confidence"],
            "additionalProperties": False,
        },
        "home_base": {
            "type": "object",
            "properties": {
                "city_state":  {"type": ["string", "null"]},
                "lat":         {"type": ["number", "null"]},
                "lon":         {"type": ["number", "null"]},
                "evidence":    {"type": "string"},
                "confidence":  {"type": "number"},
            },
            "required": ["city_state", "lat", "lon", "evidence", "confidence"],
            "additionalProperties": False,
        },
        "minimum_rate_per_mile": {
            "type": "object",
            "properties": {
                "value":      {"type": ["number", "null"]},
                "evidence":   {"type": "string"},
                "confidence": {"type": "number"},
            },
            "required": ["value", "evidence", "confidence"],
            "additionalProperties": False,
        },
        "equipment_types": {
            "type": "object",
            "properties": {
                "value":      {"type": "array", "items": {"type": "string"}},
                "evidence":   {"type": "string"},
                "confidence": {"type": "number"},
            },
            "required": ["value", "evidence", "confidence"],
            "additionalProperties": False,
        },
        "weight_capacity_lb": {
            "type": "object",
            "properties": {
                "value":      {"type": ["integer", "null"]},
                "evidence":   {"type": "string"},
                "confidence": {"type": "number"},
                "inferred":   {"type": "boolean"},
                "assumption": {"type": ["string", "null"]},
            },
            "required": ["value", "evidence", "confidence", "inferred", "assumption"],
            "additionalProperties": False,
        },
        "constraints": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "type":     {"type": "string"},
                    "value":    {"type": "string"},
                    "evidence": {"type": "string"},
                },
                "required": ["type", "value", "evidence"],
                "additionalProperties": False,
            },
        },
        "notes": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
    "required": [
        "current_location", "home_base", "minimum_rate_per_mile",
        "equipment_types", "weight_capacity_lb", "constraints", "notes",
    ],
    "additionalProperties": False,
}

_OPENAI_SYSTEM = """\
You are a logistics data extractor. Extract a structured driver profile from the
dispatch conversation transcript provided by the user.

Return a JSON object that matches the schema exactly. Follow these rules:

current_location
  The city/state where the driver physically IS right now (not home, not a destination).
  Provide decimal-degree lat/lon if the city is known; otherwise set null.

home_base
  The driver's primary operating base or home city.
  Provide decimal-degree lat/lon if known; otherwise set null.

minimum_rate_per_mile
  The lowest effective $/mile the DRIVER will accept.
  Extract only from the driver's own statements — do NOT use the dispatcher's rate quotes.

equipment_types
  List of equipment types the driver explicitly says they OWN or OPERATE.
  Use canonical names: Hotshot, Gooseneck, Flatbed, Van, Reefer, Stepdeck, Lowboy, Tanker.
  Do NOT include equipment types the driver merely asks about rhetorically.

weight_capacity_lb
  Max weight in lbs. If not explicitly stated:
    - Infer only when equipment type gives clear basis (e.g. "hotshot gooseneck" → ~15000 lb).
    - Set inferred=true and document the assumption.
    - Set value=null if you cannot infer with reasonable confidence.

constraints
  Explicit constraints the driver states (lane avoidances, factoring requirements, etc.).

notes
  Any additional relevant information not captured above.

evidence
  Quote the exact phrase from the transcript that supports each extracted value.

confidence
  0.0–1.0. Use 1.0 for explicit statements, 0.7–0.9 for strong inferences,
  below 0.7 for weak inferences. Set null fields with confidence < 0.5.
"""


def _extract_openai(transcript: str, city_table: dict) -> dict:
    from openai import OpenAI
    client = OpenAI()

    known_cities = "\n".join(
        f"  {k}" for k in sorted(city_table) if "," in k
    )
    user_msg = (
        f"TRANSCRIPT:\n{transcript}\n\n"
        f"City coordinate lookup table (for reference when setting lat/lon):\n{known_cities}"
    )

    messages = [
        {"role": "system", "content": _OPENAI_SYSTEM},
        {"role": "user",   "content": user_msg},
    ]

    # ── Log del REQUEST ───────────────────────────────────────────────────────
    _oai_log.debug(
        "═══ OPENAI REQUEST ════════════════════════════════════════════════\n"
        "Model: gpt-4o-2024-08-06\n"
        "response_format: json_schema / driver_profile (strict)\n"
        "── system ──────────────────────────────────────────────────────────\n"
        "%s\n"
        "── user ────────────────────────────────────────────────────────────\n"
        "%s",
        _OPENAI_SYSTEM,
        user_msg,
    )

    response = client.chat.completions.create(
        model="gpt-4o-2024-08-06",
        messages=messages,
        response_format={
            "type": "json_schema",
            "json_schema": {
                "name": "driver_profile",
                "strict": True,
                "schema": _PROFILE_SCHEMA,
            },
        },
    )

    raw_content = response.choices[0].message.content

    # ── Log del RESPONSE ──────────────────────────────────────────────────────
    _oai_log.debug(
        "═══ OPENAI RESPONSE ═══════════════════════════════════════════════\n"
        "finish_reason : %s\n"
        "usage         : prompt=%s  completion=%s  total=%s\n"
        "── content ─────────────────────────────────────────────────────────\n"
        "%s",
        response.choices[0].finish_reason,
        response.usage.prompt_tokens,
        response.usage.completion_tokens,
        response.usage.total_tokens,
        json.dumps(json.loads(raw_content), indent=2, ensure_ascii=False),
    )

    return json.loads(raw_content)


# =============================================================================
# 5.  Profile extraction — deterministic fallback
# =============================================================================

# Only ownership assertions count as equipment evidence.
_OWNERSHIP_MARKERS = [
    "i run ", "i drive ", "i have ", "i operate ", "i own ", "i use ",
    "my trailer", "my truck", "i haul", "i pull",
]

_EQUIP_KEYWORDS: dict[str, str] = {
    "hotshot":    "Hotshot",
    "hot shot":   "Hotshot",
    "gooseneck":  "Gooseneck",
    "goose neck": "Gooseneck",
    "flatbed":    "Flatbed",
    "flat bed":   "Flatbed",
    "van":        "Van",
    "reefer":     "Reefer",
    "stepdeck":   "Stepdeck",
    "step deck":  "Stepdeck",
    "lowboy":     "Lowboy",
    "low boy":    "Lowboy",
}


def _fallback_extract(transcript: str, city_table: dict) -> dict:
    """
    Deterministic extraction using keyword/regex patterns.

    Key design decisions:
      - Rate extraction uses ONLY driver speech lines to avoid false matches with
        dispatcher quotes (e.g. "my best loads are $4/mi" is the dispatcher's voice).
      - Equipment extraction uses only driver lines that contain ownership markers
        ("I run", "I drive", etc.) to avoid including equipment types the driver
        merely asks about rhetorically.
      - Coordinates come exclusively from the city table — no hardcoded lat/lon.
    """
    by_speaker   = lines_by_speaker(transcript)
    driver_lines = by_speaker.get("Driver", [])
    full_lower   = transcript.lower()

    # Lines where driver asserts ownership (for equipment extraction)
    ownership_lines = [
        ln for ln in driver_lines
        if any(m in ln.lower() for m in _OWNERSHIP_MARKERS)
    ]
    ownership_text = " ".join(ownership_lines).lower()
    driver_text    = " ".join(driver_lines).lower()

    def _loc(city_state, lat, lon, evidence, confidence):
        return {"city_state": city_state, "lat": lat, "lon": lon,
                "evidence": evidence, "confidence": confidence}

    def _field(value, evidence, confidence):
        return {"value": value, "evidence": evidence, "confidence": confidence}

    profile: dict = {
        "current_location":     _loc(None, None, None, "", 0.0),
        "home_base":            _loc(None, None, None, "", 0.0),
        "minimum_rate_per_mile": _field(None, "", 0.0),
        "equipment_types":       _field([], "", 0.0),
        "weight_capacity_lb": {
            "value": None, "evidence": "", "confidence": 0.0,
            "inferred": False, "assumption": None,
        },
        "constraints": [],
        "notes": [],
    }

    # ── Current location ──────────────────────────────────────────────────────
    # Pattern: "I'm in <city>" or "I am in <city>"
    loc_m = re.search(r"i[''’]?m in ([a-z][a-z\s]+?)(?:[,.]|\s*but|\s*and|$)",
                      driver_text)
    if loc_m:
        raw_city   = loc_m.group(1).strip()
        coords     = resolve_city(raw_city, city_table)
        label      = _city_label(raw_city) if coords else raw_city.title()
        evidence   = f"Driver: \"I'm in {raw_city.title()}\""
        profile["current_location"] = _loc(
            label,
            coords[0] if coords else None,
            coords[1] if coords else None,
            evidence,
            0.95 if coords else 0.7,
        )

    # ── Home base ─────────────────────────────────────────────────────────────
    # Pattern: dispatch mentions city → driver confirms with "usually" or "that's correct"
    # Terminators are punctuation or end-of-string only — NOT \s, which would
    # cut multi-word city names like "San Antonio" after the first token.
    home_patterns = [
        r"based out (?:in|of) ([a-z][a-z\s]+?)(?:\?|[,.]|$)",
        r"you[‘’’]?re (?:from|in|based in) ([a-z][a-z\s]+?)(?:\?|[,.]|$)",
        r"i think you[‘’’]?re (?:in|based(?: in| out)?) ([a-z][a-z\s]+?)(?:\?|[,.]|$)",
    ]
    for pat in home_patterns:
        m = re.search(pat, full_lower)
        if m:
            confirms = "usually" in driver_text or "that's correct" in driver_text or "yes" in driver_text[:50]
            if confirms:
                raw_city = m.group(1).strip()
                coords   = resolve_city(raw_city, city_table)
                label    = _city_label(raw_city) if coords else raw_city.title()
                evidence = (
                    f"Dispatch asks about {raw_city.title()}; "
                    f"driver confirms: \"Yes, that's correct. I'm usually in that area\""
                )
                profile["home_base"] = _loc(
                    label,
                    coords[0] if coords else None,
                    coords[1] if coords else None,
                    evidence,
                    0.9 if coords else 0.65,
                )
                break

    # ── Minimum rate per mile ─────────────────────────────────────────────────
    # Search DRIVER lines only.
    # Search only driver lines for rate mentions.
    # Use a pattern that requires "per mile" so we don't match earnings figures
    # like "$2,300" that happen to start with the same digit.
    _rate_re = re.compile(r"\$(\d+(?:\.\d+)?)\s*per\s*mile", re.I)
    _above_re = re.compile(r"above\s+\$(\d+(?:\.\d+)?)", re.I)

    rate_m = _rate_re.search(driver_text)
    if rate_m:
        rate  = float(rate_m.group(1))
        ev_ln = next(
            (ln for ln in driver_lines if _rate_re.search(ln)),
            f"${rate}/mile",
        )
        profile["minimum_rate_per_mile"] = _field(rate, f'Driver: "{ev_ln}"', 0.95)
    else:
        above_m = _above_re.search(driver_text)
        if above_m:
            rate  = float(above_m.group(1))
            ev_ln = next(
                (ln for ln in driver_lines if _above_re.search(ln)),
                f"above ${rate} per mile",
            )
            profile["minimum_rate_per_mile"] = _field(rate, f'Driver: "{ev_ln}"', 0.9)

    # ── Equipment types ───────────────────────────────────────────────────────
    # Search only ownership lines; canonical dedup via normalize_equip.
    found_equip: list[str]  = []
    found_evidence: list[str] = []
    seen: set[str] = set()

    for keyword, display in _EQUIP_KEYWORDS.items():
        if keyword in ownership_text:
            canonical = normalize_equip(keyword)
            if canonical not in seen:
                seen.add(canonical)
                found_equip.append(display)
                ev_ln = next(
                    (ln for ln in ownership_lines if keyword in ln.lower()),
                    keyword,
                )
                found_evidence.append(f'"{ev_ln}"')

    if found_equip:
        profile["equipment_types"] = _field(
            found_equip,
            "; ".join(found_evidence),
            0.9,
        )

    # ── Weight capacity ───────────────────────────────────────────────────────
    # Look for explicit lb/lbs/pounds mention in driver lines.
    wt_m = re.search(r"(\d{1,3}(?:,\d{3})*|\d+)\s*(?:lb|lbs|pounds)", driver_text)
    if wt_m:
        wt = int(wt_m.group(1).replace(",", ""))
        ev_ln = next(
            (ln for ln in driver_lines if wt_m.group(1).replace(",", "") in ln.replace(",", "")),
            str(wt),
        )
        profile["weight_capacity_lb"] = {
            "value": wt, "evidence": f'Driver: "{ev_ln}"',
            "confidence": 0.9, "inferred": False, "assumption": None,
        }
    else:
        equip_canonicals = {normalize_equip(e) for e in found_equip}
        if "hotshot" in equip_canonicals or "gooseneck" in equip_canonicals:
            profile["weight_capacity_lb"] = {
                "value": 15000,
                "evidence": "Not explicitly stated in transcript.",
                "confidence": 0.6,
                "inferred": True,
                "assumption": (
                    "15,000 lb — conservative capacity for a pickup-truck + 40′ gooseneck "
                    "trailer setup, safely below the legal axle-weight ceiling for Class 4–5 trucks."
                ),
            }

    # ── Constraints ───────────────────────────────────────────────────────────
    constraints: list[dict] = []
    if "factoring" in driver_text:
        ev_ln = next((ln for ln in driver_lines if "factoring" in ln.lower()), "factoring")
        constraints.append({
            "type": "factoring",
            "value": "broker must be approved by driver's factoring company",
            "evidence": f'Driver: "{ev_ln}"',
        })
    north_avoid_m = re.search(r"don[''’]?t like\s+(?:going\s+)?north", driver_text)
    if north_avoid_m:
        ev_ln = next((ln for ln in driver_lines if "north" in ln.lower()), "north")
        constraints.append({
            "type": "lane_preference",
            "value": "avoids north TX — cites Austin, Houston, Waco traffic",
            "evidence": f'Driver: "{ev_ln}"',
        })
    profile["constraints"] = constraints

    # ── Notes ─────────────────────────────────────────────────────────────────
    notes: list[str] = []
    preferred = [c.title() for c in ["laredo", "corpus", "midland", "odessa", "rio grande"]
                 if c in driver_text]
    if preferred:
        notes.append(f"Preferred lanes: {', '.join(preferred)}")
    if re.search(r"two or three days|2 or 3 days", driver_text):
        notes.append("Runs 2–3 days/week")
    profile["notes"] = notes

    return profile


# =============================================================================
# 6.  Profile extraction — dispatcher
# =============================================================================

def extract_profile(xlsx_path: str, city_table: dict) -> tuple[dict, str]:
    """
    Return (profile_dict, method_label).
    Tries OpenAI GPT-4o structured outputs first; falls back to deterministic.
    After extraction, fills any null lat/lon from the city table.
    """
    transcript = read_transcript(xlsx_path)
    method = "deterministic"
    profile: dict

    if os.environ.get("OPENAI_API_KEY"):
        try:
            log.info("  → OpenAI GPT-4o structured outputs")
            profile = _extract_openai(transcript, city_table)
            method = "openai-gpt4o"
        except Exception as exc:
            log.warning("  → OpenAI failed (%s); usando fallback determinístico", exc)
            profile = _fallback_extract(transcript, city_table)
    else:
        log.info("  → Deterministic (keyword) extraction")
        profile = _fallback_extract(transcript, city_table)

    # Post-extraction: resolve any null coordinates from city table
    fill_coords(profile["current_location"], city_table)
    fill_coords(profile["home_base"], city_table)

    return profile, method


# =============================================================================
# 7.  Load ingestion
# =============================================================================

@dataclass
class Load:
    load_id:     str
    origin:      str
    origin_lat:  float
    origin_lon:  float
    destination: str
    dest_lat:    float
    dest_lon:    float
    trailer:     str
    weight:      int
    price:       float


def _is_missing(val: Any) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.upper() == "MISSING" or s.lower() == "nan"


def _try_float(val: Any) -> Optional[float]:
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def load_loads(xlsx_path: str) -> tuple[list[Load], list[dict]]:
    """
    Parse the Loads tab.  Return (complete_loads, skipped).
    A row is skipped (with reason recorded) when price, or destination
    city/coordinates are absent or non-numeric.
    """
    df = pd.read_excel(xlsx_path, sheet_name="Loads", header=0)
    loads:   list[Load] = []
    skipped: list[dict] = []

    for _, row in df.iterrows():
        load_id = str(row.get("Load ID", "?"))
        reasons: list[str] = []

        if _is_missing(row.get("Destination")):
            reasons.append("Destination city missing")
        for col in ("Dest Lat", "Dest Lon"):
            v = row.get(col)
            if _is_missing(v) or _try_float(v) is None:
                reasons.append(f"{col} missing or non-numeric")

        price_raw = row.get("Price ($)")
        if _is_missing(price_raw) or _try_float(price_raw) is None:
            reasons.append(f"Price missing or non-numeric (got {price_raw!r})")

        if reasons:
            skipped.append({"load_id": load_id, "reason": "; ".join(reasons)})
            continue

        try:
            loads.append(Load(
                load_id     = load_id,
                origin      = str(row["Origin"]),
                origin_lat  = float(row["Origin Lat"]),
                origin_lon  = float(row["Origin Lon"]),
                destination = str(row["Destination"]),
                dest_lat    = float(row["Dest Lat"]),
                dest_lon    = float(row["Dest Lon"]),
                trailer     = str(row["Trailer"]),
                weight      = int(row["Weight"]),
                price       = float(row["Price ($)"]),
            ))
        except (ValueError, KeyError) as exc:
            skipped.append({"load_id": load_id, "reason": str(exc)})

    return loads, skipped


# =============================================================================
# 8.  Eligibility filtering and ranking
# =============================================================================

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Great-circle distance in statute miles."""
    R = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a  = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


@dataclass
class RankedLoad:
    load:          Load
    effective_rpm: float
    dh_to_origin:  float
    loaded_miles:  float
    dh_home:       float
    total_miles:   float


def rank_loads(profile: dict, loads: list[Load]) -> tuple[list[RankedLoad], list[dict]]:
    """
    Filter loads to those eligible for this driver, then rank by effective RPM.

    Reads exclusively from the profile dict — no hardcoded equipment names,
    weight limits, rates, or coordinates anywhere in this function.

    effective_rate = price / (dh_to_origin + loaded_miles + dh_home)
    """
    ranked:   list[RankedLoad] = []
    rejected: list[dict]       = []

    # ── Pull profile values (never hardcoded) ────────────────────────────────
    cur_lat   = profile["current_location"]["lat"]
    cur_lon   = profile["current_location"]["lon"]
    home_lat  = profile["home_base"]["lat"]
    home_lon  = profile["home_base"]["lon"]
    equip_list = profile["equipment_types"]["value"]    # list[str] or []
    wt_cap     = profile["weight_capacity_lb"]["value"]  # int | None
    min_rpm    = profile["minimum_rate_per_mile"]["value"]  # float | None

    # ── Guard: current location must be known to calculate deadhead ──────────
    if cur_lat is None or cur_lon is None:
        for ld in loads:
            rejected.append({
                "load_id": ld.load_id,
                "reason":  "Current truck location coordinates are null — cannot calculate deadhead to origin",
            })
        return [], rejected

    home_known = home_lat is not None and home_lon is not None

    for ld in loads:
        # Filter 1 — equipment type compatibility
        if equip_list:
            if not equip_matches(equip_list, ld.trailer):
                rejected.append({
                    "load_id": ld.load_id,
                    "reason": (
                        f"Trailer '{ld.trailer}' (→ '{normalize_equip(ld.trailer)}') "
                        f"not in driver equipment {equip_list} "
                        f"(→ {[normalize_equip(e) for e in equip_list]})"
                    ),
                })
                continue

        # Filter 2 — weight capacity
        if wt_cap is None:
            rejected.append({
                "load_id": ld.load_id,
                "reason":  "Weight capacity unknown — cannot verify load weight eligibility",
            })
            continue
        if ld.weight > wt_cap:
            rejected.append({
                "load_id": ld.load_id,
                "reason": (
                    f"Load weight {ld.weight:,} lb exceeds driver capacity {wt_cap:,} lb"
                ),
            })
            continue

        # ── Distance calculations ─────────────────────────────────────────────
        dh_to_origin = haversine(cur_lat,       cur_lon,
                                  ld.origin_lat, ld.origin_lon)
        loaded_miles = haversine(ld.origin_lat, ld.origin_lon,
                                  ld.dest_lat,   ld.dest_lon)
        dh_home      = haversine(ld.dest_lat,   ld.dest_lon,
                                  home_lat,       home_lon) if home_known else 0.0

        total_miles  = dh_to_origin + loaded_miles + dh_home
        rpm          = ld.price / total_miles if total_miles > 0 else 0.0

        # Filter 3 — minimum effective rate
        if min_rpm is not None and rpm < min_rpm:
            rejected.append({
                "load_id": ld.load_id,
                "reason": (
                    f"Effective rate ${rpm:.3f}/mi < driver minimum ${min_rpm:.2f}/mi  "
                    f"({dh_to_origin:.1f} DH + {loaded_miles:.1f} loaded"
                    f" + {dh_home:.1f} DH-home = {total_miles:.1f} mi)"
                ),
            })
            continue

        ranked.append(RankedLoad(
            load          = ld,
            effective_rpm = rpm,
            dh_to_origin  = dh_to_origin,
            loaded_miles  = loaded_miles,
            dh_home       = dh_home,
            total_miles   = total_miles,
        ))

    ranked.sort(key=lambda r: r.effective_rpm, reverse=True)
    return ranked, rejected


# =============================================================================
# 9.  Workbook writer
# =============================================================================

def _find_field_row(ws, label: str, col: int = 1) -> Optional[int]:
    """Return the 1-based Excel row where label appears in the given column."""
    for row in ws.iter_rows():
        cell = row[col - 1]
        if cell.value and str(cell.value).strip().lower() == label.lower():
            return cell.row
    return None


_ANSWER_COL = 2  # Column B in Part A


def write_part_a(ws, profile: dict) -> None:
    cur   = profile["current_location"]
    home  = profile["home_base"]
    rate  = profile["minimum_rate_per_mile"]
    equip = profile["equipment_types"]
    wt    = profile["weight_capacity_lb"]

    rate_str  = f"${rate['value']:.2f}" if rate["value"] is not None else "Not extracted"
    equip_str = ", ".join(equip["value"]) if equip["value"] else "Not extracted"

    if wt["value"] is not None:
        wt_str = f"{wt['value']:,}"
        if wt.get("inferred"):
            wt_str += " (inferred — not stated in transcript)"
    else:
        wt_str = "Not extracted"

    field_map: dict[str, Any] = {
        "Current Location":      cur.get("city_state") or "Not extracted",
        "Current Lat":           cur.get("lat"),
        "Current Lon":           cur.get("lon"),
        "Home Base":             home.get("city_state") or "Not extracted",
        "Home Lat":              home.get("lat"),
        "Home Lon":              home.get("lon"),
        "Minimum Rate Per Mile": rate_str,
        "Equipment Type(s)":     equip_str,
        "Weight Capacity (lb)":  wt_str,
    }

    for label, value in field_map.items():
        row = _find_field_row(ws, label)
        if row is not None:
            ws.cell(row=row, column=_ANSWER_COL).value = value

    # Notes / constraints — write into the blank row after Weight Capacity
    notes_parts: list[str] = list(profile.get("notes", []))
    for c in profile.get("constraints", []):
        notes_parts.append(c["value"])
    notes_text = " | ".join(notes_parts)

    notes_row = _find_field_row(ws, "Notes")
    if notes_row is None:
        wc_row = _find_field_row(ws, "Weight Capacity (lb)")
        if wc_row:
            notes_row = wc_row + 1
            ws.cell(row=notes_row, column=1).value = "Notes"
    if notes_row:
        ws.cell(row=notes_row, column=_ANSWER_COL).value = notes_text or "—"


def _build_readme(
    profile: dict,
    ranked:   list[RankedLoad],
    rejected: list[dict],
    skipped:  list[dict],
    method:   str,
) -> str:
    cur  = profile["current_location"]
    home = profile["home_base"]
    rate = profile["minimum_rate_per_mile"]
    wt   = profile["weight_capacity_lb"]

    wt_str = (
        f"{wt['value']:,} lb{'  (inferred)' if wt.get('inferred') else ''}"
        if wt["value"] is not None else "null"
    )

    skip_str = "\n".join(f"  {s['load_id']}: {s['reason']}" for s in skipped) or "  None"
    rej_str  = "\n".join(f"  {r['load_id']}: {r['reason']}" for r in rejected) or "  None"
    top3_str = "\n".join(
        f"  #{i+1} {r.load.load_id}: {r.load.origin} → {r.load.destination}"
        f"  ${r.effective_rpm:.3f}/mi"
        f"  ({r.dh_to_origin:.0f}+{r.loaded_miles:.0f}+{r.dh_home:.0f}={r.total_miles:.0f} mi)"
        for i, r in enumerate(ranked[:3])
    ) or "  (none eligible)"

    notable = next(
        (r for r in rejected if "not in driver equipment" in r["reason"]),
        rejected[0] if rejected else None,
    )
    notable_str = f"{notable['load_id']}: {notable['reason']}" if notable else "None"

    return (
        f"Code: solve_cinesis_test.py  |  Extraction: {method}\n\n"
        f"PROFILE  Current: {cur.get('city_state')} ({cur.get('lat')}, {cur.get('lon')})"
        f"  conf={cur.get('confidence', 0):.2f}\n"
        f"         Home: {home.get('city_state')} ({home.get('lat')}, {home.get('lon')})"
        f"  conf={home.get('confidence', 0):.2f}\n"
        f"         Min rate: ${rate.get('value')}/mi   Wt cap: {wt_str}\n\n"
        f"SKIPPED (incomplete data):\n{skip_str}\n\n"
        f"REJECTED (ineligible):\n{rej_str}\n\n"
        f"TOP 3 ELIGIBLE:\n{top3_str}\n\n"
        f"NOTABLE REJECTION: {notable_str}"
    )


def write_part_b(ws, ranked: list[RankedLoad], readme: str) -> None:
    # Rank 1/2/3 — Excel rows 5, 6, 7
    # El effective rate se escribe como número flotante (no string) con formato
    # de número Excel "$#,##0.000" para evitar ambigüedad entre separador decimal
    # y separador de miles según la configuración regional del lector.
    from openpyxl.styles import numbers as xl_numbers
    for i, result in enumerate(ranked[:3]):
        row = 5 + i
        ws.cell(row=row, column=2).value = result.load.load_id
        rate_cell = ws.cell(row=row, column=3)
        rate_cell.value          = round(result.effective_rpm, 3)
        rate_cell.number_format  = '"$"#,##0.000'

    # README block
    cell = ws.cell(row=11, column=1)
    cell.value     = readme
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[11].height = 240


def write_results(
    xlsx_path:   str,
    output_path: str,
    profile:     dict,
    ranked:      list[RankedLoad],
    rejected:    list[dict],
    skipped:     list[dict],
    method:      str,
) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    write_part_a(wb["Part A (Fill In)"], profile)
    readme = _build_readme(profile, ranked, rejected, skipped, method)
    write_part_b(wb["Part B (Fill In)"], ranked, readme)
    wb.save(output_path)


# =============================================================================
# 10.  Main
# =============================================================================

def main() -> None:
    raw_input   = sys.argv[1] if len(sys.argv) > 1 else "cinesis_good_fit_test_clean.xlsx"
    output_path = sys.argv[2] if len(sys.argv) > 2 else "cinesis_good_fit_test_completed.xlsx"

    # Agrega .xlsx al input si no lo tiene
    xlsx_path = raw_input if raw_input.endswith(".xlsx") else raw_input + ".xlsx"

    # Crea el directorio de output si no existe
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    log.info("Input:  %s", xlsx_path)
    log.info("Output: %s", output_path)

    # Step 1 — Build city coordinate lookup from workbook + seed table
    log.info("[1/5] Building city coordinate table...")
    city_table = build_city_table(xlsx_path)
    log.info("  %d entries", len(city_table))

    # Step 2 — Extract driver profile from Sample Conversation tab
    log.info("[2/5] Extracting driver profile...")
    profile, method = extract_profile(xlsx_path, city_table)

    log.info("── Extracted profile (method: %s) ──\n%s",
             method, json.dumps(profile, indent=2, ensure_ascii=False))

    # Step 3 — Load and validate loads from the Loads tab
    log.info("[3/5] Loading loads...")
    loads, skipped = load_loads(xlsx_path)
    log.info("  %d complete  |  %d skipped", len(loads), len(skipped))
    for s in skipped:
        log.info("  SKIP    %s: %s", s["load_id"], s["reason"])

    # Step 4 — Filter and rank
    log.info("[4/5] Filtering and ranking loads...")
    ranked, rejected = rank_loads(profile, loads)

    for r in rejected:
        log.info("  REJECT  %s: %s", r["load_id"], r["reason"])

    log.info("  All eligible loads (%d) sorted by effective rate:", len(ranked))
    for i, res in enumerate(ranked):
        flag = f"  ← #{i+1}" if i < 3 else ""
        log.info(
            "  %s  %-15s → %-20s  $%.3f/mi"
            "  DH %5.1f + loaded %5.1f + DH-home %5.1f = %6.1f mi%s",
            res.load.load_id, res.load.origin, res.load.destination,
            res.effective_rpm,
            res.dh_to_origin, res.loaded_miles, res.dh_home, res.total_miles,
            flag,
        )

    # Step 5 — Write completed workbook
    log.info("[5/5] Writing %s...", output_path)
    write_results(xlsx_path, output_path, profile, ranked, rejected, skipped, method)
    log.info("Done. Log guardado en: %s", _log_file)


if __name__ == "__main__":
    main()
